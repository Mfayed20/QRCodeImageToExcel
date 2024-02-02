from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from urllib.parse import urlparse
import re

class ExcelWorkbookCreator:
    def __init__(self, successful_extractions, failed_extractions, output_excel):
        self.successful_extractions = successful_extractions
        self.failed_extractions = failed_extractions
        self.output_excel = output_excel

    def create_workbook(self):
        wb = Workbook()
        ws_success = wb.create_sheet("Successful Extractions", 0)
        self.populate_sheet(ws_success, self.successful_extractions, 'successful')
        ws_failed = wb.create_sheet("Failed Extractions", 1)
        self.populate_sheet(ws_failed, self.failed_extractions, 'failed')
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
        wb.save(self.output_excel)

    def populate_sheet(self, ws, data, extraction_type):
        column_names = ['Image Name', 'URL', 'Date', 'Status'] if extraction_type == 'successful' else ['Image Name']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color=Color("FFFFFF"))
        ws.append(column_names)
        for row in data:
            ws.append([f'=HYPERLINK("{row.get(col, "")}", "{self.simplify_url(row.get(col, ""))}")' if col == 'URL' else row.get(col, '') for col in column_names])
            for i, cell in enumerate(ws[1]):
                cell.alignment = Alignment(horizontal="left")
                cell.fill = header_fill
                cell.font = header_font
        self.apply_styles_and_validation(ws, column_names)
        self.adjust_column_widths(ws, column_names)

    def simplify_url(self, url):
        if not url:
            return ''
        parsed_url = urlparse(url)
        simplified_url = parsed_url.netloc + parsed_url.path
        return re.sub(r'/+$', '', simplified_url)  # Remove trailing slashes

    def apply_styles_and_validation(self, ws, column_names):
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        date_style.alignment = Alignment(horizontal='left')
        alignment = Alignment(horizontal='left')
        if 'Date' in column_names:
            date_column = column_names.index('Date') + 1
            for row in ws.iter_rows(min_row=2, min_col=date_column, max_col=date_column):
                for cell in row:
                    cell.style = date_style
                    cell.alignment = alignment
        if 'Status' in column_names:
            status_col_letter = get_column_letter(column_names.index('Status') + 1)
            dv = DataValidation(type="list", formula1='"Applied, Approved, Rejected, Not Applicable"')
            ws.add_data_validation(dv)
            for row in ws[f"{status_col_letter}2:{status_col_letter}{ws.max_row}"]:
                for cell in row:
                    dv.add(cell)
        self.apply_conditional_formatting(ws, column_names)

    def apply_conditional_formatting(self, ws, column_names):
        fills = {
            'Applied': PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid"),
            'Approved': PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            'Rejected': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            'Not Applicable': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }
        if 'Status' in column_names:
            status_col_letter = get_column_letter(column_names.index('Status') + 1)
            for status, fill in fills.items():
                rule = Rule(type="expression", dxf=DifferentialStyle(fill=fill))
                rule.formula = [f'INDIRECT("{status_col_letter}"&ROW())="{status}"']
                ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}{ws.max_row}', rule)

    def adjust_column_widths(self, ws, column_names):
        for i, column in enumerate(ws.columns):
            column = [cell for cell in column]
            if column_names[i] == 'URL':
                hyperlink_lengths = [
                    len(re.findall(r'"(.+?)"', str(cell.value))[1])
                    for cell in column if "HYPERLINK" in str(cell.value)
                ]
                max_length = max(hyperlink_lengths) if hyperlink_lengths else 15
            elif column_names[i] == 'Status':
                max_length = 15
            else:
                max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 1)
            ws.column_dimensions[get_column_letter(
                column[0].column)].width = adjusted_width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
